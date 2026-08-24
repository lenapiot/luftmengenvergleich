from __future__ import annotations

import argparse
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# KONFIGURATION
# ============================================================

# Beispiele:
#   L1 1.1
#   L1 10.38
#   L1 14.40
#
# Leerzeichen innerhalb der PDF dürfen variieren.
SCHEMA_NUMBER_PATTERN = re.compile(
    r"""
    (?<![A-Z0-9])
    L\s*(\d{1,2})
    \s+
    (\d{1,3})
    \s*\.\s*
    (\d{1,3})
    (?![\d.])
    """,
    re.IGNORECASE | re.VERBOSE,
)

BML_HEADER_NAMES = {
    "ep_schemanummer",
    "epschemanummer",
    "ep schemanummer",
}

PREFERRED_BML_SHEET = "Komponentenliste Unternehmer"


# ============================================================
# DATENKLASSEN
# ============================================================

@dataclass
class SchemaRecord:
    schemanummer: str
    quelle: str
    seite: int
    farbe: str


@dataclass
class BMLRecord:
    schemanummer: str
    quelle: str
    tabellenblatt: str
    zeile: int


# ============================================================
# NORMALISIERUNG
# ============================================================

def normalize_schema_number(value: object) -> str | None:
    """
    Normalisiert eine Lüftungs-Schemanummer.

    Beispiele:
        "L1 1.1"     -> "L1 1.1"
        "L 1 14. 40" -> "L1 14.40"
    """
    if value is None:
        return None

    text = (
        str(value)
        .replace("\u00a0", " ")
        .replace("\n", " ")
        .replace("\t", " ")
        .strip()
        .upper()
    )

    match = SCHEMA_NUMBER_PATTERN.search(text)
    if match is None:
        return None

    system = int(match.group(1))
    group = int(match.group(2))
    number = int(match.group(3))

    return f"L{system} {group}.{number}"


def schema_number_sort_key(value: str) -> tuple[int, int, int, str]:
    match = re.fullmatch(
        r"L(\d+)\s+(\d+)\.(\d+)",
        value,
        re.IGNORECASE,
    )
    if match is None:
        return (999999, 999999, 999999, value)

    return (
        int(match.group(1)),
        int(match.group(2)),
        int(match.group(3)),
        value,
    )


# ============================================================
# FARBEN IM PDF
# ============================================================

def rgb_from_int(color: int) -> tuple[int, int, int]:
    return (
        (color >> 16) & 255,
        (color >> 8) & 255,
        color & 255,
    )


def classify_luft_blue(color: int) -> str | None:
    """
    Akzeptiert die beiden relevanten Lüftungsfarben:

    - Hellblau / Cyan
    - Dunkelblau

    Die Grenzen sind absichtlich etwas tolerant, damit kleine
    Farbabweichungen durch PDF-Export nicht zum Verlust von Nummern führen.
    """
    red, green, blue = rgb_from_int(color)

    # Hellblau / Cyan:
    # typischer Testplan: RGB (0, 255, 255)
    if (
        blue >= 170
        and green >= 150
        and red <= 100
    ):
        return "Hellblau"

    # Dunkelblau:
    # typischer Testplan: RGB (0, 0, 255)
    if (
        blue >= 140
        and green <= 130
        and red <= 130
    ):
        return "Dunkelblau"

    return None


# ============================================================
# PDF-EXTRAKTION
# ============================================================

def extract_schema_numbers(
    schema_pdf: str | Path,
) -> list[SchemaRecord]:
    """
    Extrahiert Lüftungs-Schemanummern aus hellblauen UND dunkelblauen
    Textspans eines Prinzipschemas.
    """
    schema_pdf = Path(schema_pdf)

    if not schema_pdf.exists():
        raise FileNotFoundError(
            f"Prinzipschema nicht gefunden: {schema_pdf}"
        )

    records: list[SchemaRecord] = []

    document = fitz.open(schema_pdf)

    try:
        for page_index, page in enumerate(document, start=1):
            page_dict = page.get_text("dict")

            for block in page_dict.get("blocks", []):
                for line in block.get("lines", []):
                    spans = line.get("spans", [])

                    # 1) Normalfall:
                    # Jede farbige Textspan einzeln prüfen.
                    line_match_count = 0

                    for span in spans:
                        text = str(
                            span.get("text", "")
                        )

                        if not text.strip():
                            continue

                        color_name = classify_luft_blue(
                            int(span.get("color", 0))
                        )

                        if color_name is None:
                            continue

                        for match in SCHEMA_NUMBER_PATTERN.finditer(text):
                            number = normalize_schema_number(
                                match.group(0)
                            )

                            if number is None:
                                continue

                            records.append(
                                SchemaRecord(
                                    schemanummer=number,
                                    quelle=schema_pdf.name,
                                    seite=page_index,
                                    farbe=color_name,
                                )
                            )
                            line_match_count += 1

                    # 2) Fallback:
                    # Falls eine Nummer durch mehrere farbige Spans
                    # getrennt wurde, die relevanten Spans der Zeile
                    # zusammensetzen.
                    if line_match_count == 0:
                        colored_parts: list[str] = []
                        colors: set[str] = set()

                        for span in spans:
                            text = str(
                                span.get("text", "")
                            )

                            if not text.strip():
                                continue

                            color_name = classify_luft_blue(
                                int(span.get("color", 0))
                            )

                            if color_name is None:
                                continue

                            colored_parts.append(text)
                            colors.add(color_name)

                        if not colored_parts:
                            continue

                        combined_text = " ".join(
                            colored_parts
                        )

                        for match in SCHEMA_NUMBER_PATTERN.finditer(
                            combined_text
                        ):
                            number = normalize_schema_number(
                                match.group(0)
                            )

                            if number is None:
                                continue

                            if len(colors) == 1:
                                color_name = next(
                                    iter(colors)
                                )
                            else:
                                color_name = "Gemischt"

                            records.append(
                                SchemaRecord(
                                    schemanummer=number,
                                    quelle=schema_pdf.name,
                                    seite=page_index,
                                    farbe=color_name,
                                )
                            )

    finally:
        document.close()

    return records


# ============================================================
# BML-EXTRAKTION
# ============================================================

def normalize_header(value: object) -> str:
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace("\u00a0", " ")
        .replace("_", "_")
    )


def find_bml_schema_column(
    workbook,
) -> tuple[str, int, int]:
    """
    Sucht die Spalte 'ep_Schemanummer'.

    Bevorzugt wird das Tabellenblatt
    'Komponentenliste Unternehmer'.

    Wichtig:
    Die Suche erfolgt zeilenweise mit iter_rows().
    Das ist bei grossen BML-Dateien wesentlich schneller als
    wiederholte cell()-Zugriffe, insbesondere im read_only-Modus.

    Rückgabe:
        (Tabellenblatt, Kopfzeile, Spaltennummer)
    """
    candidates: list[
        tuple[int, str, int, int]
    ] = []

    for worksheet in workbook.worksheets:
        max_search_row = min(
            worksheet.max_row,
            20,
        )

        for row_index, values in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=max_search_row,
                values_only=True,
            ),
            start=1,
        ):
            for column_index, value in enumerate(
                values,
                start=1,
            ):
                normalized = normalize_header(
                    value
                )

                compact = (
                    normalized
                    .replace("_", "")
                    .replace(" ", "")
                )

                is_header = (
                    normalized in BML_HEADER_NAMES
                    or compact == "epschemanummer"
                )

                if not is_header:
                    continue

                priority = (
                    0
                    if worksheet.title
                    == PREFERRED_BML_SHEET
                    else 1
                )

                candidates.append(
                    (
                        priority,
                        worksheet.title,
                        row_index,
                        column_index,
                    )
                )

    if not candidates:
        raise ValueError(
            "Die Spalte 'ep_Schemanummer' wurde "
            "in der BML nicht gefunden."
        )

    candidates.sort(
        key=lambda item: (
            item[0],
            item[1],
            item[2],
            item[3],
        )
    )

    _, sheet_name, header_row, column = (
        candidates[0]
    )

    return (
        sheet_name,
        header_row,
        column,
    )


def extract_bml_numbers(
    bml_excel: str | Path,
) -> list[BMLRecord]:
    """
    Liest die Schemanummern aus der Spalte 'ep_Schemanummer'
    der BML.

    Die Datenzeilen werden nur einmal sequenziell gelesen.
    Dadurch bleibt die Auswertung auch bei grossen BML-Dateien schnell.
    """
    bml_excel = Path(bml_excel)

    if not bml_excel.exists():
        raise FileNotFoundError(
            f"BML nicht gefunden: {bml_excel}"
        )

    workbook = load_workbook(
        bml_excel,
        read_only=True,
        data_only=True,
    )

    try:
        (
            sheet_name,
            header_row,
            schema_column,
        ) = find_bml_schema_column(
            workbook
        )

        worksheet = workbook[
            sheet_name
        ]

        records: list[BMLRecord] = []

        for row_index, row_values in enumerate(
            worksheet.iter_rows(
                min_row=header_row + 1,
                min_col=schema_column,
                max_col=schema_column,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            value = (
                row_values[0]
                if row_values
                else None
            )

            number = normalize_schema_number(
                value
            )

            if number is None:
                continue

            records.append(
                BMLRecord(
                    schemanummer=number,
                    quelle=bml_excel.name,
                    tabellenblatt=sheet_name,
                    zeile=row_index,
                )
            )

        return records

    finally:
        workbook.close()


# ============================================================
# VERGLEICH
# ============================================================

def determine_status(
    schema_count: int,
    bml_count: int,
) -> str:
    if schema_count > 1 and bml_count > 1:
        return "Mehrfach in beiden"

    if schema_count > 1:
        return "Mehrfach im Schema"

    if bml_count > 1:
        return "Mehrfach in BML"

    if schema_count == 1 and bml_count == 1:
        return "OK"

    if schema_count == 1 and bml_count == 0:
        return "Nur im Schema"

    if schema_count == 0 and bml_count == 1:
        return "Nur in BML"

    return "Unklar"


def compare_numbers(
    schema_records: Iterable[SchemaRecord],
    bml_records: Iterable[BMLRecord],
) -> list[dict[str, object]]:
    schema_counter = Counter(
        record.schemanummer
        for record in schema_records
    )

    bml_counter = Counter(
        record.schemanummer
        for record in bml_records
    )

    all_numbers = sorted(
        set(schema_counter)
        | set(bml_counter),
        key=schema_number_sort_key,
    )

    result: list[
        dict[str, object]
    ] = []

    for number in all_numbers:
        schema_count = schema_counter.get(
            number,
            0,
        )
        bml_count = bml_counter.get(
            number,
            0,
        )

        result.append(
            {
                "Schemanummer": number,
                "Status": determine_status(
                    schema_count,
                    bml_count,
                ),
                "Anzahl Schema": schema_count,
                "Anzahl BML": bml_count,
            }
        )

    status_order = {
        "Nur im Schema": 0,
        "Nur in BML": 1,
        "Mehrfach im Schema": 2,
        "Mehrfach in BML": 3,
        "Mehrfach in beiden": 4,
        "Unklar": 5,
        "OK": 6,
    }

    result.sort(
        key=lambda row: (
            status_order.get(
                str(row["Status"]),
                99,
            ),
            schema_number_sort_key(
                str(row["Schemanummer"])
            ),
        )
    )

    return result


# ============================================================
# EXCEL-AUSGABE
# ============================================================

def style_worksheet(
    worksheet,
) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value)),
            )

        column_letter = get_column_letter(
            column_cells[0].column
        )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                max_length + 2,
                12,
            ),
            45,
        )


def export_result(
    comparison: list[dict[str, object]],
    schema_records: list[SchemaRecord],
    bml_records: list[BMLRecord],
    output_path: str | Path,
) -> Path:
    """
    Erstellt die Excel-Auswertung direkt mit openpyxl.

    Dadurch existiert von Anfang an mindestens ein sichtbares
    Tabellenblatt und der Export ist auch für die spätere EXE robust.
    """
    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    # --------------------------------------------------------
    # VERGLEICH
    # --------------------------------------------------------

    comparison_sheet = workbook.active
    comparison_sheet.title = "Vergleich"

    comparison_sheet.append(
        [
            "Schemanummer",
            "Status",
            "Anzahl Schema",
            "Anzahl BML",
        ]
    )

    for row in comparison:
        comparison_sheet.append(
            [
                row["Schemanummer"],
                row["Status"],
                row["Anzahl Schema"],
                row["Anzahl BML"],
            ]
        )

    # --------------------------------------------------------
    # NUMMERN SCHEMA
    # --------------------------------------------------------

    schema_sheet = workbook.create_sheet(
        "Nummern_Schema"
    )

    schema_sheet.append(
        [
            "Schemanummer",
            "Farbe",
            "Seite",
            "Quelldatei",
        ]
    )

    for record in sorted(
        schema_records,
        key=lambda item: (
            schema_number_sort_key(
                item.schemanummer
            ),
            item.seite,
            item.farbe,
        ),
    ):
        schema_sheet.append(
            [
                record.schemanummer,
                record.farbe,
                record.seite,
                record.quelle,
            ]
        )

    # --------------------------------------------------------
    # NUMMERN BML
    # --------------------------------------------------------

    bml_sheet = workbook.create_sheet(
        "Nummern_BML"
    )

    bml_sheet.append(
        [
            "Schemanummer",
            "Tabellenblatt",
            "Zeile",
            "Quelldatei",
        ]
    )

    for record in sorted(
        bml_records,
        key=lambda item: (
            schema_number_sort_key(
                item.schemanummer
            ),
            item.zeile,
        ),
    ):
        bml_sheet.append(
            [
                record.schemanummer,
                record.tabellenblatt,
                record.zeile,
                record.quelle,
            ]
        )

    # --------------------------------------------------------
    # FORMATIERUNG
    # --------------------------------------------------------

    status_colors = {
        "OK": "C6EFCE",
        "Nur im Schema": "FFF2CC",
        "Nur in BML": "F4CCCC",
        "Mehrfach im Schema": "FCE4D6",
        "Mehrfach in BML": "FCE4D6",
        "Mehrfach in beiden": "EADCF8",
        "Unklar": "D9EAD3",
    }

    for worksheet in workbook.worksheets:
        style_worksheet(
            worksheet
        )

    for row in range(
        2,
        comparison_sheet.max_row + 1,
    ):
        status = comparison_sheet.cell(
            row=row,
            column=2,
        ).value

        color = status_colors.get(
            status
        )

        if color is None:
            continue

        for column in range(
            1,
            comparison_sheet.max_column + 1,
        ):
            comparison_sheet.cell(
                row=row,
                column=column,
            ).fill = PatternFill(
                fill_type="solid",
                fgColor=color,
            )

    workbook.save(
        output_path
    )
    workbook.close()

    return output_path


# ============================================================
# HAUPTFUNKTION
# ============================================================

def run_luft_number_check(
    schema_pdf: str | Path,
    bml_excel: str | Path,
    output_dir: str | Path,
    name: str = "Lueftung_Schemanummernkontrolle",
) -> Path:
    """
    Führt die komplette Lüftungs-Schemanummernkontrolle aus.
    """
    schema_pdf = Path(
        schema_pdf
    )
    bml_excel = Path(
        bml_excel
    )
    output_dir = Path(
        output_dir
    )

    if schema_pdf.suffix.lower() != ".pdf":
        raise ValueError(
            "Das Prinzipschema muss eine PDF-Datei sein."
        )

    if bml_excel.suffix.lower() not in {
        ".xlsx",
        ".xlsm",
    }:
        raise ValueError(
            "Die BML muss eine Excel-Datei (.xlsx oder .xlsm) sein."
        )

    schema_records = extract_schema_numbers(
        schema_pdf
    )

    if not schema_records:
        raise ValueError(
            "Im Prinzipschema wurden keine hellblauen oder "
            "dunkelblauen Lüftungs-Schemanummern erkannt."
        )

    bml_records = extract_bml_numbers(
        bml_excel
    )

    if not bml_records:
        raise ValueError(
            "In der BML wurden in der Spalte "
            "'ep_Schemanummer' keine gültigen Schemanummern erkannt."
        )

    comparison = compare_numbers(
        schema_records,
        bml_records,
    )

    safe_name = re.sub(
        r'[<>:"/\\|?*]+',
        "_",
        str(name).strip(),
    )

    if not safe_name:
        safe_name = (
            "Lueftung_Schemanummernkontrolle"
        )

    output_path = output_dir / (
        f"{safe_name}.xlsx"
    )

    return export_result(
        comparison=comparison,
        schema_records=schema_records,
        bml_records=bml_records,
        output_path=output_path,
    )


# ============================================================
# TEST / DIREKTER AUFRUF
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Vergleicht Lüftungs-Schemanummern "
            "zwischen Prinzipschema und BML."
        )
    )

    parser.add_argument(
        "schema_pdf",
        help="Pfad zum Lüftungs-Prinzipschema (PDF)",
    )
    parser.add_argument(
        "bml_excel",
        help="Pfad zur Betriebsmittelliste (XLSX)",
    )
    parser.add_argument(
        "output_dir",
        help="Ausgabeordner",
    )
    parser.add_argument(
        "--name",
        default="Lueftung_Schemanummernkontrolle",
        help="Name der Ausgabedatei ohne .xlsx",
    )

    args = parser.parse_args()

    output_path = run_luft_number_check(
        schema_pdf=args.schema_pdf,
        bml_excel=args.bml_excel,
        output_dir=args.output_dir,
        name=args.name,
    )

    print(
        f"Auswertung erstellt: {output_path}"
    )


if __name__ == "__main__":
    main()
