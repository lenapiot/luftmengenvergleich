from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from config.settings import EXCEL_STATUS_COLORS

def find_column_number(
   worksheet,
   column_name: str,
) -> int | None:
   """Sucht eine Spalte anhand ihrer Überschrift."""
   for cell in worksheet[1]:
       if cell.value == column_name:
           return cell.column
   return None

def adjust_column_widths(
   worksheet,
) -> None:
   """Passt die Spaltenbreiten automatisch an."""
   for column_cells in worksheet.columns:
       max_length = 0
       column_number = column_cells[0].column
       column_letter = get_column_letter(
           column_number
       )
       for cell in column_cells:
           if cell.value is not None:
               max_length = max(
                   max_length,
                   len(str(cell.value)),
               )
       worksheet.column_dimensions[
           column_letter
       ].width = min(
           max_length + 2,
           45,
       )

def format_data_sheet(
   worksheet,
) -> None:
   """Formatiert ein normales Datenblatt."""
   header_fill = PatternFill(
       fill_type="solid",
       fgColor="D9E1F2",
   )
   thin_border = Border(
       left=Side(
           style="thin",
           color="D9D9D9",
       ),
       right=Side(
           style="thin",
           color="D9D9D9",
       ),
       top=Side(
           style="thin",
           color="D9D9D9",
       ),
       bottom=Side(
           style="thin",
           color="D9D9D9",
       ),
   )
   for cell in worksheet[1]:
       cell.font = Font(
           bold=True
       )
       cell.fill = header_fill
       cell.alignment = Alignment(
           horizontal="center",
           vertical="center",
           wrap_text=True,
       )
       cell.border = thin_border
   worksheet.freeze_panes = "A2"
   worksheet.auto_filter.ref = (
       worksheet.dimensions
   )
   for row in worksheet.iter_rows(
       min_row=2,
       max_row=worksheet.max_row,
   ):
       for cell in row:
           cell.alignment = Alignment(
               vertical="top",
               wrap_text=True,
           )
           cell.border = thin_border
   adjust_column_widths(
       worksheet
   )

def color_rows_by_status(
   worksheet,
) -> None:
   """Färbt Zeilen anhand der Statusspalte."""
   status_column = find_column_number(
       worksheet,
       "status",
   )
   if status_column is None:
       return
   for row_number in range(
       2,
       worksheet.max_row + 1,
   ):
       status_cell = worksheet.cell(
           row=row_number,
           column=status_column,
       )
       status = status_cell.value
       color = EXCEL_STATUS_COLORS.get(
           status
       )
       if color is None:
           continue
       fill = PatternFill(
           fill_type="solid",
           fgColor=color,
       )
       for cell in worksheet[row_number]:
           cell.fill = fill
       status_cell.font = Font(
           bold=True
       )

def create_legend_sheet(
   workbook,
   floorplan_pdfs: list[Path],
   schema_pdfs: list[Path],
   comparison_df: pd.DataFrame,
) -> None:
   """Erstellt Legende und Statuszusammenfassung."""
   if "Legende" in workbook.sheetnames:
       del workbook["Legende"]
   worksheet = workbook.create_sheet(
       "Legende",
       0,
   )
   worksheet["A1"] = "Luftmengen-Vergleich"
   worksheet["A1"].font = Font(
       bold=True,
       size=16,
   )
   worksheet["A2"] = (
       "Automatischer Vergleich zwischen "
       "mehreren Luftmengen-Grundrissen und "
       "Lüftungs-Prinzipschemata"
   )
   worksheet["A4"] = "Status"
   worksheet["B4"] = "Bedeutung"
   worksheet["C4"] = "Anzahl"
   meanings = {
       "OK": (
           "Raum ist in Grundriss und Schema "
           "vorhanden; Raumname und Luftmengen "
           "stimmen überein."
       ),
       "Abweichung Luftmenge": (
           "Zuluft und/oder Abluft "
           "unterscheiden sich."
       ),
       "Abweichung Raumname": (
           "Die Luftmengen stimmen, aber "
           "die Raumbezeichnung weicht ab."
       ),
       "Nur im Grundriss": (
           "Der Raum wurde nur in einem "
           "Grundriss gefunden."
       ),
       "Nur im Schema": (
           "Der Raum wurde nur in einem "
           "Prinzipschema gefunden."
       ),
       "Mehrfach / uneindeutig": (
           "Eine Raumnummer wurde mehrfach "
           "mit widersprüchlichen Angaben "
           "gefunden."
       ),
   }
   counts = (
       comparison_df["status"]
       .astype(str)
       .value_counts()
       .to_dict()
   )
   row_number = 5
   for status, meaning in meanings.items():
       worksheet.cell(
           row=row_number,
           column=1,
           value=status,
       )
       worksheet.cell(
           row=row_number,
           column=2,
           value=meaning,
       )
       worksheet.cell(
           row=row_number,
           column=3,
           value=counts.get(
               status,
               0,
           ),
       )
       worksheet.cell(
           row=row_number,
           column=1,
       ).fill = PatternFill(
           fill_type="solid",
           fgColor=EXCEL_STATUS_COLORS[
               status
           ],
       )
       row_number += 1
   worksheet["A13"] = (
       "Verwendete Grundrisse:"
   )
   worksheet["B13"] = "\n".join(
       floorplan_pdf.name
       for floorplan_pdf in floorplan_pdfs
   )
   worksheet["A14"] = (
       "Verwendete Schemata:"
   )
   worksheet["B14"] = "\n".join(
       schema_pdf.name
       for schema_pdf in schema_pdfs
   )
   worksheet["A16"] = "Hinweis:"
   worksheet["B16"] = (
       "Die Auswertung ist für ähnlich "
       "strukturierte, digital erzeugte "
       "PDFs ausgelegt. Die Ergebnisse "
       "müssen fachlich kontrolliert werden."
   )
   for reference in [
       "A13",
       "A14",
       "A16",
   ]:
       worksheet[
           reference
       ].font = Font(
           bold=True
       )
   thin_border = Border(
       left=Side(
           style="thin",
           color="B7B7B7",
       ),
       right=Side(
           style="thin",
           color="B7B7B7",
       ),
       top=Side(
           style="thin",
           color="B7B7B7",
       ),
       bottom=Side(
           style="thin",
           color="B7B7B7",
       ),
   )
   for row in worksheet.iter_rows(
       min_row=4,
       max_row=10,
       min_col=1,
       max_col=3,
   ):
       for cell in row:
           cell.border = thin_border
           cell.alignment = Alignment(
               vertical="top",
               wrap_text=True,
           )
   for cell in worksheet[4]:
       cell.font = Font(
           bold=True
       )
       cell.fill = PatternFill(
           fill_type="solid",
           fgColor="D9E1F2",
       )
   for reference in [
       "B13",
       "B14",
       "B16",
   ]:
       worksheet[
           reference
       ].alignment = Alignment(
           vertical="top",
           wrap_text=True,
       )
   worksheet.column_dimensions[
       "A"
   ].width = 28
   worksheet.column_dimensions[
       "B"
   ].width = 80
   worksheet.column_dimensions[
       "C"
   ].width = 12
   worksheet.row_dimensions[13].height = max(
       18,
       15 * len(floorplan_pdfs),
   )
   worksheet.row_dimensions[14].height = max(
       18,
       15 * len(schema_pdfs),
   )
   worksheet.freeze_panes = "A5"

def export_excel(
   output_path: Path,
   floorplan_raw_df: pd.DataFrame,
   schema_raw_df: pd.DataFrame,
   comparison_df: pd.DataFrame,
   marking_df: pd.DataFrame,
   floorplan_pdfs: list[Path],
   schema_pdfs: list[Path],
) -> None:
   """Exportiert und formatiert die Excel-Auswertung."""
   deviations_df = comparison_df[
       comparison_df[
           "status"
       ].astype(str) != "OK"
   ].copy()
   only_floorplan_df = comparison_df[
       comparison_df[
           "status"
       ].astype(str) == "Nur im Grundriss"
   ].copy()
   only_schema_df = comparison_df[
       comparison_df[
           "status"
       ].astype(str) == "Nur im Schema"
   ].copy()
   with pd.ExcelWriter(
       output_path,
       engine="openpyxl",
   ) as writer:
       floorplan_raw_df.to_excel(
           writer,
           sheet_name="Grundriss_Rohdaten",
           index=False,
       )
       schema_raw_df.to_excel(
           writer,
           sheet_name="Schema_Rohdaten",
           index=False,
       )
       comparison_df.to_excel(
           writer,
           sheet_name="Vergleich",
           index=False,
       )
       deviations_df.to_excel(
           writer,
           sheet_name="Abweichungen",
           index=False,
       )
       only_floorplan_df.to_excel(
           writer,
           sheet_name="Nur_im_Grundriss",
           index=False,
       )
       only_schema_df.to_excel(
           writer,
           sheet_name="Nur_im_Schema",
           index=False,
       )
       marking_df.to_excel(
           writer,
           sheet_name="PDF_Markierungen",
           index=False,
       )
   workbook = load_workbook(
       output_path
   )
   create_legend_sheet(
       workbook,
       floorplan_pdfs,
       schema_pdfs,
       comparison_df,
   )
   for worksheet in workbook.worksheets:
       if worksheet.title == "Legende":
           continue
       format_data_sheet(
           worksheet
       )
       if worksheet.title in {
           "Vergleich",
           "Abweichungen",
           "Nur_im_Grundriss",
           "Nur_im_Schema",
       }:
           color_rows_by_status(
               worksheet
           )
   workbook.save(
       output_path
   )