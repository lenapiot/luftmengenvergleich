"""
Automatischer Luftmengenvergleich für ähnlich aufgebaute Grundriss- und
Lüftungsplan-PDFs.

Das Programm:
1. lässt Grundriss und Schema auswählen,
2. liest alle PDF-Seiten aus,
3. extrahiert Raumnummer, Raumname, Zuluft und Abluft,
4. vergleicht beide Dokumente,
5. erstellt eine formatierte Excel-Datei,
6. markiert die Räume im Grundriss-PDF,
7. erstellt ein Protokoll nicht eindeutiger Markierungen.

Geeignet für digital erzeugte PDFs mit ähnlicher Textstruktur wie das Testpaar:
    Grundriss:
        Raumname / Raumnummer
        ZUL: 4'000 m³/h
        ABL: 4'450 m³/h

    Schema:
        Raumname
        Raumnummer
        Zuluft 4'000 m³/h
        Abluft 4'450 m³/h

Installation:
    python -m pip install pymupdf pandas openpyxl

Start:
    python luftmengen_vergleich.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import fitz
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


import re

from config.settings import(
    ROOM_RE,
    MAX_AIRFLOW_DISTANCE,
    ROOM_SEARCH_RADIUS,
    EXCEL_STATUS_COLORS,
    PDF_STATUS_COLORS,
)


from ui.dialogs import(
    choose_pdf,
    choose_output_folder,
    validate_pdf,
)

from export.pdf_export import create_marked_pdf

from export.excel_export import export_excel

from core.comparison import build_comparison

from core.extraction import(
    extract_clean_lines, 
    split_lines_by_page,
    extract_rooms_from_pages)

# =============================================================================
# 1. EINSTELLUNGEN (settings.py)
# =============================================================================

# =============================================================================
# 2. DATENSTRUKTUR (extraction.py)
# =============================================================================

# =============================================================================
# 3. ALLGEMEINE HILFSFUNKTIONEN (extraction.py)
# ===========================================================================

# =============================================================================
# 4. PDF EINLESEN (extraction.py)
# =============================================================================

# =============================================================================
# 5. RÄUME UND LUFTMENGEN EXTRAHIEREN (extraction.py)
# =============================================================================

# =============================================================================
# 6. DOPPELTE ODER UNEINDEUTIGE RÄUME KONSOLIDIEREN
# =============================================================================



# =============================================================================
# 7. GRUNDRISS UND SCHEMA VERGLEICHEN
# =============================================================================



# =============================================================================
# 8. PDF-MARKIERUNG (pdf_export.py)
# =============================================================================


# =============================================================================
# 9. EXCEL-DATEI ERSTELLEN UND FORMATIEREN (excel_export.py)
# =============================================================================

# =============================================================================
# 10. DATEIEN AUSWÄHLEN (dialogs.py)
# =============================================================================


# =============================================================================
# 11. HAUPTPROGRAMM
# =============================================================================

def main() -> None:
    """Führt den vollständigen Luftmengenvergleich aus."""
    print()
    print("Luftmengen-Vergleich")
    print("====================")
    print()

    floorplan_pdf = choose_pdf(
        "Grundriss-PDF auswählen"
    )

    schema_pdf = choose_pdf(
        "Lüftungsplan / Prinzipschema auswählen"
    )

    validate_pdf(
        floorplan_pdf,
        "Grundriss",
    )

    validate_pdf(
        schema_pdf,
        "Schema",
    )

    output_dir = choose_output_folder(
        floorplan_pdf.parent
    )

    output_excel = (
        output_dir
        / (
            f"{floorplan_pdf.stem}"
            "_Luftmengenvergleich.xlsx"
        )
    )

    output_pdf = (
        output_dir
        / (
            f"{floorplan_pdf.stem}"
            "_markiert.pdf"
        )
    )

    print("1/6 PDFs werden eingelesen ...")

    floorplan_records = (
        extract_clean_lines(
            floorplan_pdf
        )
    )

    schema_records = (
        extract_clean_lines(
            schema_pdf
        )
    )

    print("2/6 Räume werden extrahiert ...")

    floorplan_raw_df = (
        extract_rooms_from_pages(
            split_lines_by_page(
                floorplan_records
            ),
            "grundriss",
        )
    )

    schema_raw_df = (
        extract_rooms_from_pages(
            split_lines_by_page(
                schema_records
            ),
            "schema",
        )
    )

    print(
        "   Grundriss:",
        len(floorplan_raw_df),
        "gefundene Datensätze",
    )

    print(
        "   Schema:",
        len(schema_raw_df),
        "gefundene Datensätze",
    )

    if floorplan_raw_df.empty:
        raise RuntimeError(
            "Im Grundriss wurden keine Räume "
            "mit ZUL/ABL erkannt."
        )

    if schema_raw_df.empty:
        raise RuntimeError(
            "Im Schema wurden keine Räume "
            "mit Zuluft/Abluft erkannt."
        )

    print("3/6 Daten werden verglichen ...")

    comparison_df = build_comparison(
        floorplan_raw_df,
        schema_raw_df,
    )

    print("4/6 Grundriss wird markiert ...")

    marking_df = create_marked_pdf(
        floorplan_pdf,
        output_pdf,
        comparison_df,
    )

    print(
        "5/6 Excel-Auswertung "
        "wird erstellt ..."
    )

    export_excel(
        output_excel,
        floorplan_raw_df,
        schema_raw_df,
        comparison_df,
        marking_df,
        floorplan_pdf,
        schema_pdf,
    )

    print("6/6 Fertig.")
    print()

    print("Excel-Auswertung:")
    print(output_excel)
    print()

    print("Markierter Grundriss:")
    print(output_pdf)
    print()

    print("Statusübersicht:")
    print(
        comparison_df[
            "status"
        ]
        .astype(str)
        .value_counts()
        .to_string()
    )


if __name__ == "__main__":
    main()
