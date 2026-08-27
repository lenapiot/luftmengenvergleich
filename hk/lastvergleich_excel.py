from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# FARBEN / STYLE
# ============================================================

DARK_BLUE = "17365D"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"

GREEN = "C6EFCE"
GREEN_TEXT = "006100"

RED = "FFC7CE"
RED_TEXT = "9C0006"

ORANGE = "FCE4D6"
ORANGE_TEXT = "9C5700"

YELLOW = "FFF2CC"
YELLOW_TEXT = "7F6000"

GREY = "E7E6E6"
DARK_GREY = "595959"

WHITE = "FFFFFF"
BLACK = "000000"

THIN_GREY = Side(
    style="thin",
    color="D9D9D9",
)


# ============================================================
# HILFSFUNKTIONEN
# ============================================================

def _safe_text(value) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value)


def _safe_number(value):
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    return value


def _join_file_names(
    paths: Iterable[str | Path],
) -> str:
    return "\n".join(
        Path(path).name
        for path in paths
    )


def _status_fill_and_font(
    status: str,
) -> tuple[PatternFill, Font]:
    normalized = (
        status
        .strip()
        .casefold()
    )

    if normalized == "ok":
        return (
            PatternFill(
                "solid",
                fgColor=GREEN,
            ),
            Font(
                color=GREEN_TEXT,
                bold=True,
            ),
        )

    if normalized == "abweichung":
        return (
            PatternFill(
                "solid",
                fgColor=RED,
            ),
            Font(
                color=RED_TEXT,
                bold=True,
            ),
        )

    if normalized == "unvollständig":
        return (
            PatternFill(
                "solid",
                fgColor=ORANGE,
            ),
            Font(
                color=ORANGE_TEXT,
                bold=True,
            ),
        )

    if (
        "mehrfach" in normalized
        or normalized == "prüfen"
    ):
        return (
            PatternFill(
                "solid",
                fgColor=YELLOW,
            ),
            Font(
                color=YELLOW_TEXT,
                bold=True,
            ),
        )

    if normalized == "keine leistung":
        return (
            PatternFill(
                "solid",
                fgColor=GREY,
            ),
            Font(
                color=DARK_GREY,
            ),
        )

    return (
        PatternFill(
            fill_type=None,
        ),
        Font(
            color=BLACK,
        ),
    )


def _style_title(cell) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=DARK_BLUE,
    )
    cell.font = Font(
        color=WHITE,
        bold=True,
        size=14,
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )


def _style_section_header(cell) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=BLUE,
    )
    cell.font = Font(
        bold=True,
        color=BLACK,
    )
    cell.alignment = Alignment(
        vertical="center",
    )


def _style_table_header(cell) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=DARK_BLUE,
    )
    cell.font = Font(
        color=WHITE,
        bold=True,
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = Border(
        bottom=THIN_GREY,
    )


def _autofit_with_limits(
    worksheet,
    min_width: int = 10,
    max_width: int = 40,
) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(
            column_cells[0].column
        )

        longest = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            text = str(value)

            line_length = max(
                (
                    len(line)
                    for line in text.splitlines()
                ),
                default=0,
            )

            longest = max(
                longest,
                line_length,
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                longest + 2,
                min_width,
            ),
            max_width,
        )


# ============================================================
# ÜBERSICHT
# ============================================================

def _write_overview_sheet(
    workbook: Workbook,
    comparison: pd.DataFrame,
    schema_pdf: str | Path,
    heating_pdfs: Iterable[str | Path],
    cooling_pdfs: Iterable[str | Path],
    building: str,
    source_type: str = "pdf",
    comparison_scope: str = "beides",
) -> None:
    ws = workbook.active
    ws.title = "Übersicht"

    scope = comparison.attrs.get(
        "vergleichsumfang",
        {},
    )

    considered_levels = scope.get(
        "beruecksichtigte_ebenen",
        [],
    )

    ignored_levels = scope.get(
        "ausgeschlossene_schema_ebenen",
        [],
    )

    note = scope.get(
        "hinweis",
        (
            "Verglichen werden nur die Ebenen, "
            "für die Heizlast- oder Kühllast-Grundrisse "
            "ausgewählt wurden."
        ),
    )

    ws.merge_cells(
        "A1:F1"
    )

    ws["A1"] = (
        "Lastvergleich Lastquelle ↔ Strangschema"
    )

    _style_title(
        ws["A1"]
    )

    ws.row_dimensions[
        1
    ].height = 24

    # --------------------------------------------------------
    # VERGLEICHSUMFANG
    # --------------------------------------------------------

    ws.merge_cells(
        "A3:F3"
    )

    ws["A3"] = (
        "Vergleichsumfang"
    )

    _style_section_header(
        ws["A3"]
    )

    not_checked = comparison.attrs.get(
        "nicht_gepruefte_raeume"
    )

    not_checked_count = (
        int(
            not_checked["raumnummer"].nunique()
        )
        if (
            isinstance(
                not_checked,
                pd.DataFrame,
            )
            and not not_checked.empty
            and "raumnummer" in not_checked.columns
        )
        else 0
    )

    source_label = (
        "Excel"
        if source_type == "excel"
        else "PDF-Grundrisse"
    )

    scope_label = {
        "heizung": "nur Heizlast",
        "kuehlung": "nur Kühllast",
        "beides": "Heiz- und Kühllast",
    }.get(
        comparison_scope,
        comparison_scope,
    )


    source_note = (
        "Bei mehreren Excel-Dateien bleibt pro Raum sichtbar, in welcher "
        "Datei bzw. in welchen Dateien er gefunden wurde. Unterschiedliche "
        "Werte für denselben Raum werden als Mehrfachfall zur Prüfung markiert."
        if source_type == "excel"
        else
        "Bei mehreren PDF-Dateien bleibt pro Raum die jeweilige Funddatei sichtbar."
    )

    overview_rows = [
        (
            "Gebäude",
            building,
        ),
        (
            "Lastquelle",
            source_label,
        ),
        (
            "Prüfumfang",
            scope_label,
        ),
        (
            "Hinweis Quelldateien",
            source_note,
        ),
        (
            "Räume anderes Gebäude (nicht geprüft)",
            not_checked_count,
        ),
        (
            "WICHTIGER HINWEIS",
            (
                note
                + " Räume eines anderen Gebäudeteils als das jeweilige "
                + "Strangschema werden im Blatt «Nicht geprüft» dokumentiert "
                + "und nicht als Fehler gewertet."
            ),
        ),
    ]

    start_row = 4

    for index, (
        label,
        value,
    ) in enumerate(
        overview_rows,
        start=start_row,
    ):
        ws.cell(
            index,
            1,
            label,
        )

        ws.cell(
            index,
            2,
            value,
        )

        ws.merge_cells(
            start_row=index,
            start_column=2,
            end_row=index,
            end_column=6,
        )

        ws.cell(
            index,
            1,
        ).font = Font(
            bold=True,
        )

        ws.cell(
            index,
            2,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    ws[
        "A8"
    ].fill = PatternFill(
        "solid",
        fgColor=YELLOW,
    )

    ws[
        "A8"
    ].font = Font(
        bold=True,
        color=YELLOW_TEXT,
    )

    ws[
        "B8"
    ].fill = PatternFill(
        "solid",
        fgColor=YELLOW,
    )

    ws[
        "B8"
    ].font = Font(
        bold=True,
        color=YELLOW_TEXT,
    )

    ws.row_dimensions[
        8
    ].height = 45

    # --------------------------------------------------------
    # DATEIEN
    # --------------------------------------------------------

    ws.merge_cells(
        "A10:F10"
    )

    ws["A10"] = (
        "Verwendete Dateien"
    )

    _style_section_header(
        ws["A10"]
    )

    if source_type == "excel":
        excel_files = list(
            heating_pdfs
        ) or list(
            cooling_pdfs
        )

        file_rows = [
            (
                "Strangschema",
                Path(schema_pdf).name,
            ),
            (
                "Heiz-/Kühllast-Excel-Datei(en)",
                _join_file_names(
                    excel_files
                ),
            ),
        ]
    else:
        file_rows = [
            (
                "Strangschema",
                Path(schema_pdf).name,
            ),
        ]

        if comparison_scope in {
            "heizung",
            "beides",
        }:
            file_rows.append(
                (
                    "Heizlast-Grundrisse",
                    _join_file_names(
                        heating_pdfs
                    ),
                )
            )

        if comparison_scope in {
            "kuehlung",
            "beides",
        }:
            file_rows.append(
                (
                    "Kühllast-Grundrisse",
                    _join_file_names(
                        cooling_pdfs
                    ),
                )
            )

    for index, (
        label,
        value,
    ) in enumerate(
        file_rows,
        start=11,
    ):
        ws.cell(
            index,
            1,
            label,
        )

        ws.cell(
            index,
            1,
        ).font = Font(
            bold=True,
        )

        ws.cell(
            index,
            2,
            value,
        )

        ws.merge_cells(
            start_row=index,
            start_column=2,
            end_row=index,
            end_column=6,
        )

        ws.cell(
            index,
            2,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        # Zeilenhöhe automatisch an die Anzahl der Dateinamen anpassen,
        # damit bei mehreren Grundrissen alle Dateien sichtbar sind.
        line_count = max(
            1,
            str(value).count("\n") + 1,
        )

        ws.row_dimensions[
            index
        ].height = max(
            18,
            15 * line_count,
        )

    # --------------------------------------------------------
    # STATUSZUSAMMENFASSUNG
    # --------------------------------------------------------

    ws.merge_cells(
        "A15:F15"
    )

    ws["A15"] = (
        "Ergebnisübersicht"
    )

    _style_section_header(
        ws["A15"]
    )

    status_order = [
        "OK",
        "Abweichung",
        "Unvollständig",
        "Mehrfach / prüfen",
        "Prüfen",
        "Keine Leistung",
    ]

    counts = (
        comparison[
            "status_gesamt"
        ]
        .value_counts()
        .to_dict()
    )

    ws[
        "A16"
    ] = "Status"

    ws[
        "B16"
    ] = "Anzahl"

    _style_table_header(
        ws[
            "A16"
        ]
    )

    _style_table_header(
        ws[
            "B16"
        ]
    )

    current_row = 17

    for status in status_order:
        count = int(
            counts.get(
                status,
                0,
            )
        )

        if count == 0:
            continue

        ws.cell(
            current_row,
            1,
            status,
        )

        ws.cell(
            current_row,
            2,
            count,
        )

        fill, font = (
            _status_fill_and_font(
                status
            )
        )

        ws.cell(
            current_row,
            1,
        ).fill = fill

        ws.cell(
            current_row,
            1,
        ).font = font

        current_row += 1

    ws.cell(
        current_row + 1,
        1,
        "Vergleichte Räume",
    )

    ws.cell(
        current_row + 1,
        2,
        len(
            comparison
        ),
    )

    ws.cell(
        current_row + 2,
        1,
        "Erstellt am",
    )

    ws.cell(
        current_row + 2,
        2,
        datetime.now().strftime(
            "%d.%m.%Y %H:%M"
        ),
    )

    ws.freeze_panes = "A4"

    ws.column_dimensions[
        "A"
    ].width = 30

    for col in (
        "B",
        "C",
        "D",
        "E",
        "F",
    ):
        ws.column_dimensions[
            col
        ].width = 22


# ============================================================
# LASTVERGLEICH
# ============================================================

def _write_comparison_sheet(
    workbook: Workbook,
    comparison: pd.DataFrame,
) -> None:
    ws = workbook.create_sheet(
        "Lastvergleich"
    )

    columns = [
        (
            "Raumnummer",
            "raumnummer",
        ),
        (
            "Raumname",
            "raumname",
        ),
        (
            "Heizlast Quelle Original [W]",
            "heizlast_original_w",
        ),
        (
            "Heizlast Vergleich [W]",
            "heizlast_vergleich_w",
        ),
        (
            "Q_H Schema [W]",
            "q_h_schema_w",
        ),
        (
            "Differenz Heizung [W]",
            "differenz_heizung_w",
        ),
        (
            "Status Heizung",
            "status_heizung",
        ),
        (
            "Heizlast Marker",
            "heizlast_marker_typ",
        ),
        (
            "Kühllast Quelle Original [W]",
            "kuehllast_original_w",
        ),
        (
            "Kühllast Vergleich [W]",
            "kuehllast_vergleich_w",
        ),
        (
            "Q_K Schema [W]",
            "q_k_schema_w",
        ),
        (
            "Differenz Kühlung [W]",
            "differenz_kuehlung_w",
        ),
        (
            "Status Kühlung",
            "status_kuehlung",
        ),
        (
            "Kühllast Marker",
            "kuehllast_marker_typ",
        ),
        (
            "Schema Status",
            "schema_status",
        ),
        (
            "Schema Werte / Konflikte",
            "schema_werte",
        ),
        (
            "Gesamtstatus",
            "status_gesamt",
        ),
        (
            "Heizlast Funddatei(en)",
            "datei_heizlast",
        ),
        (
            "Anzahl Heizlast-Dateien",
            "anzahl_dateien_heizlast",
        ),
        (
            "Heizlast in mehreren Dateien",
            "mehrere_dateien_heizlast",
        ),
        (
            "Kühllast Funddatei(en)",
            "datei_kuehllast",
        ),
        (
            "Anzahl Kühllast-Dateien",
            "anzahl_dateien_kuehllast",
        ),
        (
            "Kühllast in mehreren Dateien",
            "mehrere_dateien_kuehllast",
        ),
        (
            "Schema-Datei",
            "datei_schema",
        ),
    ]

    headers = [
        item[
            0
        ]
        for item in columns
    ]

    keys = [
        item[
            1
        ]
        for item in columns
    ]

    ws.append(
        headers
    )

    for cell in ws[
        1
    ]:
        _style_table_header(
            cell
        )

    numeric_keys = {
        "heizlast_vergleich_w",
        "q_h_schema_w",
        "differenz_heizung_w",
        "kuehllast_vergleich_w",
        "q_k_schema_w",
        "differenz_kuehlung_w",
    }

    for _, row in comparison.iterrows():
        row_values = []

        for key in keys:
            value = row.get(
                key
            )

            # Rückwärtskompatibel:
            # Falls die Vergleichstabelle die neuen Zählfelder noch nicht
            # explizit enthält, werden sie aus den Funddatei-Spalten
            # abgeleitet.
            if key == "anzahl_dateien_heizlast":
                source_text = _safe_text(
                    row.get(
                        "datei_heizlast"
                    )
                )
                value = (
                    len(
                        [
                            part
                            for part in source_text.split(" | ")
                            if part.strip()
                        ]
                    )
                    if source_text
                    else 0
                )

            elif key == "mehrere_dateien_heizlast":
                source_text = _safe_text(
                    row.get(
                        "datei_heizlast"
                    )
                )
                count = (
                    len(
                        [
                            part
                            for part in source_text.split(" | ")
                            if part.strip()
                        ]
                    )
                    if source_text
                    else 0
                )
                value = (
                    "Ja"
                    if count > 1
                    else (
                        "Nein"
                        if count == 1
                        else ""
                    )
                )

            elif key == "anzahl_dateien_kuehllast":
                source_text = _safe_text(
                    row.get(
                        "datei_kuehllast"
                    )
                )
                value = (
                    len(
                        [
                            part
                            for part in source_text.split(" | ")
                            if part.strip()
                        ]
                    )
                    if source_text
                    else 0
                )

            elif key == "mehrere_dateien_kuehllast":
                source_text = _safe_text(
                    row.get(
                        "datei_kuehllast"
                    )
                )
                count = (
                    len(
                        [
                            part
                            for part in source_text.split(" | ")
                            if part.strip()
                        ]
                    )
                    if source_text
                    else 0
                )
                value = (
                    "Ja"
                    if count > 1
                    else (
                        "Nein"
                        if count == 1
                        else ""
                    )
                )

            row_values.append(
                (
                    _safe_number(
                        value
                    )
                    if key in numeric_keys
                    else _safe_text(
                        value
                    )
                )
            )

        ws.append(
            row_values
        )

    header_map = {
        cell.value:
            cell.column
        for cell in ws[
            1
        ]
    }

    status_columns = [
        "Status Heizung",
        "Status Kühlung",
        "Gesamtstatus",
    ]

    for row_index in range(
        2,
        ws.max_row + 1,
    ):
        for header in status_columns:
            column_index = (
                header_map[
                    header
                ]
            )

            cell = ws.cell(
                row_index,
                column_index,
            )

            fill, font = (
                _status_fill_and_font(
                    _safe_text(
                        cell.value
                    )
                )
            )

            cell.fill = fill
            cell.font = font

        overall_cell = ws.cell(
            row_index,
            header_map[
                "Gesamtstatus"
            ],
        )

        overall_status = _safe_text(
            overall_cell.value
        )

        if overall_status in {
            "Abweichung",
            "Unvollständig",
            "Mehrfach / prüfen",
            "Prüfen",
        }:
            for column_index in range(
                1,
                ws.max_column + 1,
            ):
                cell = ws.cell(
                    row_index,
                    column_index,
                )

                if cell.fill.fill_type is None:
                    if overall_status == "Abweichung":
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="FFF0F0",
                        )

                    elif overall_status == "Unvollständig":
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="FFF8F2",
                        )

                    else:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="FFFBEA",
                        )

        for cell in ws[
            row_index
        ]:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # Bewusst KEIN Excel-Tabellenobjekt.
    # Der normale Autofilter unten reicht aus und verhindert
    # die Excel-Reparaturmeldung beim Öffnen der Datei.

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(ws.max_column)}"
        f"{ws.max_row}"
    )

    # feste sinnvolle Breiten
    fixed_widths = {
        "A": 16,
        "B": 30,
        "C": 22,
        "D": 20,
        "E": 16,
        "F": 19,
        "G": 22,
        "H": 20,
        "I": 22,
        "J": 20,
        "K": 16,
        "L": 19,
        "M": 22,
        "N": 20,
        "O": 28,
        "P": 42,
        "Q": 20,
        "R": 38,
        "S": 38,
        "T": 38,
    }

    for column_letter, width in fixed_widths.items():
        ws.column_dimensions[
            column_letter
        ].width = width

    ws.row_dimensions[
        1
    ].height = 42


# ============================================================
# PRÜFPROTOKOLL
# ============================================================

def _write_checks_sheet(
    workbook: Workbook,
    heating_check: pd.DataFrame,
    cooling_check: pd.DataFrame,
) -> None:
    ws = workbook.create_sheet(
        "Dateiprüfung"
    )

    headers = [
        "Lastart",
        "Datei",
        "Erkanntes Gebäude",
        "Erwartetes Gebäude",
        "Akzeptiert",
        "Grund",
        "Räume gesamt",
        "Räume verwendet",
        "Räume anderes Gebäude",
    ]

    ws.append(
        headers
    )

    for cell in ws[
        1
    ]:
        _style_table_header(
            cell
        )

    def append_frame(
        dataframe: pd.DataFrame,
        load_type: str,
    ) -> None:
        if dataframe.empty:
            return

        for _, row in dataframe.iterrows():
            ws.append(
                [
                    load_type,
                    _safe_text(
                        row.get(
                            "datei"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "erkanntes_gebaeude"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "erwartetes_gebaeude"
                        )
                    ),
                    (
                        "Ja"
                        if bool(
                            row.get(
                                "akzeptiert",
                                False,
                            )
                        )
                        else "Nein"
                    ),
                    _safe_text(
                        row.get(
                            "grund"
                        )
                    ),
                    _safe_number(
                        row.get(
                            "anzahl_raeume"
                        )
                    ),
                    _safe_number(
                        row.get(
                            "anzahl_raeume_verwendet"
                        )
                    ),
                    _safe_number(
                        row.get(
                            "anzahl_raeume_anderes_gebaeude"
                        )
                    ),
                ]
            )

    append_frame(
        heating_check,
        "Heizlast",
    )

    append_frame(
        cooling_check,
        "Kühllast",
    )

    for row_index in range(
        2,
        ws.max_row + 1,
    ):
        accepted_cell = ws.cell(
            row_index,
            5,
        )

        if accepted_cell.value == "Ja":
            accepted_cell.fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )
            accepted_cell.font = Font(
                color=GREEN_TEXT,
                bold=True,
            )
        else:
            accepted_cell.fill = PatternFill(
                "solid",
                fgColor=RED,
            )
            accepted_cell.font = Font(
                color=RED_TEXT,
                bold=True,
            )

    ws.freeze_panes = "A2"

    _autofit_with_limits(
        ws,
        min_width=12,
        max_width=45,
    )


# ============================================================
# EXPORT
# ============================================================


# ============================================================
# NICHT GEPRÜFTE RÄUME AUS MIT12-GRUNDRISSEN
# ============================================================

def _write_not_checked_rooms_sheet(
    workbook: Workbook,
    comparison: pd.DataFrame,
) -> None:
    """
    Dokumentiert Räume aus der Lastquelle,
    die zum anderen Gebäudeteil gehören.

    Beispiel:
        Strangschema MIT2
        Grundriss enthält MIT1 + MIT2
        -> MIT2 wird verglichen
        -> MIT1 wird hier dokumentiert, aber NICHT als Fehler gewertet.
    """
    ws = workbook.create_sheet(
        "Nicht geprüft"
    )

    not_checked = comparison.attrs.get(
        "nicht_gepruefte_raeume"
    )

    ws.merge_cells(
        "A1:J1"
    )

    ws["A1"] = (
        "Nicht geprüfte Räume aus anderem Gebäudeteil"
    )

    _style_title(
        ws["A1"]
    )

    ws.merge_cells(
        "A3:J4"
    )

    ws["A3"] = (
        "Diese Räume wurden in der gewählten Lastquelle gefunden, gehören "
        "aber zu einem anderen Gebäudeteil als das jeweilige Strangschema. "
        "Sie werden dokumentiert, aber nicht verglichen und nicht als Fehler gewertet."
    )

    ws["A3"].fill = PatternFill(
        "solid",
        fgColor=YELLOW,
    )

    ws["A3"].font = Font(
        bold=True,
        color=YELLOW_TEXT,
    )

    ws["A3"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    headers = [
        "Lastart",
        "Raumnummer",
        "Raumname",
        "Leistung Original [W]",
        "Vergleichswert [W]",
        "Gebäude Raum",
        "Zielgebäude",
        "Ebene",
        "Datei",
        "Grund",
    ]

    ws.append(
        []
    )
    ws.append(
        headers
    )

    header_row = 6

    for cell in ws[
        header_row
    ]:
        _style_table_header(
            cell
        )

    if (
        isinstance(
            not_checked,
            pd.DataFrame,
        )
        and not not_checked.empty
    ):
        for _, row in not_checked.iterrows():
            ws.append(
                [
                    _safe_text(
                        row.get(
                            "lastart"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "raumnummer"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "raumname"
                        )
                    ),
                    _safe_number(
                        row.get(
                            "leistung_w"
                        )
                    ),
                    _safe_number(
                        row.get(
                            "vergleichswert_w"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "gebaeude"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "zielgebaeude"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "ebene"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "datei"
                        )
                    ),
                    _safe_text(
                        row.get(
                            "nicht_geprueft_grund"
                        )
                    ),
                ]
            )
    else:
        ws.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Keine Räume aus einem anderen Gebäudeteil erkannt.",
            ]
        )

    ws.freeze_panes = "A7"

    if ws.max_row >= header_row:
        ws.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(ws.max_column)}"
            f"{ws.max_row}"
        )

    widths = {
        "A": 14,
        "B": 18,
        "C": 30,
        "D": 20,
        "E": 20,
        "F": 16,
        "G": 16,
        "H": 10,
        "I": 42,
        "J": 55,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    for row_index in range(
        7,
        ws.max_row + 1,
    ):
        for cell in ws[
            row_index
        ]:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

def export_load_comparison_excel(
    output_path: str | Path,
    comparison: pd.DataFrame,
    schema_pdf: str | Path,
    heating_pdfs: Iterable[str | Path],
    cooling_pdfs: Iterable[str | Path],
    building: str,
    heating_check: pd.DataFrame | None = None,
    cooling_check: pd.DataFrame | None = None,
    source_type: str = "pdf",
    comparison_scope: str = "beides",
) -> Path:
    """
    Erstellt die Excel-Ausgabe für den Lastvergleich.

    Blätter:
        Übersicht
        Lastvergleich
        Dateiprüfung
        Nicht geprüft

    Die Übersicht zeigt gut sichtbar,
    welche Ebenen geprüft und welche
    Schema-Ebenen NICHT geprüft wurden.
    """

    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    _write_overview_sheet(
        workbook,
        comparison,
        schema_pdf,
        heating_pdfs,
        cooling_pdfs,
        building,
        source_type=source_type,
        comparison_scope=comparison_scope,
    )

    _write_comparison_sheet(
        workbook,
        comparison,
    )

    _write_checks_sheet(
        workbook,
        (
            heating_check
            if heating_check is not None
            else pd.DataFrame()
        ),
        (
            cooling_check
            if cooling_check is not None
            else pd.DataFrame()
        ),
    )

    _write_not_checked_rooms_sheet(
        workbook,
        comparison,
    )

    workbook.save(
        output_path
    )

    return output_path
