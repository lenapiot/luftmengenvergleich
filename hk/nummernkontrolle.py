from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import fitz
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# POSITIONSNUMMERN
# ============================================================
#
# Allgemeines Format nach Rückmeldung:
#
#   1–2 Buchstaben + 1–3 Zahlen
#   .
#   1–3 Zahlen
#   optional:
#   . + genau 1 Zahl
#
# Beispiele:
#   F24.527
#   F24.527.1
#   F24.527.2
#   D24.028
#   ZV70.01
#   ZV70.12
#   A1.2
#   A1.23
#   A1.234
#   AB123.1
#   AB123.123.4
#
# "STA" bleibt als bereits unterstützte Alt-Ausnahme erhalten,
# damit bestehende Schemen nicht verschlechtert werden.
#
# Das Muster erlaubt außerdem Leerzeichen innerhalb einer Nummer,
# wie sie durch PDF-Export entstehen können:
#
#   F 2 4 . 5 2 7 . 1
#   Z V 7 0 . 0 1
#
# Diese werden anschließend zu F24.527.1 bzw. ZV70.01 normalisiert.
# ============================================================

POSITION_PATTERN = re.compile(
    r"""
    (?<![A-Z0-9])
    (?:
        STA
        |
        [A-Z]\s*(?:[A-Z]\s*)?
    )
    (?:\d\s*){1,3}
    \.\s*
    (?:\d\s*){1,3}
    (?:
        \.\s*\d
    )?
    (?!\s*\.\s*\d)
    (?![A-Z0-9])
    """,
    re.IGNORECASE | re.VERBOSE,
)


@dataclass
class PositionRecord:
    positionsnummer: str
    quelle: str
    datei: str
    seite: int | None = None


# ============================================================
# NORMALISIERUNG / SUCHE
# ============================================================

def normalize_position_number(
    value: object,
) -> str | None:
    """
    Normalisiert eine Positionsnummer.

    Beispiele:
        F8 0 . 2 1       -> F80.21
        F 2 4 . 5 2 7.1 -> F24.527.1
        Z V 7 0 . 0 1   -> ZV70.01
    """
    if value is None:
        return None

    text = str(
        value
    ).strip()

    if not text:
        return None

    text = (
        text.replace(
            "\u00a0",
            " ",
        )
        .replace(
            "\n",
            " ",
        )
        .replace(
            "\t",
            " ",
        )
        .upper()
    )

    match = POSITION_PATTERN.search(
        text
    )

    if not match:
        return None

    number = match.group(
        0
    )

    number = re.sub(
        r"\s+",
        "",
        number,
    )

    # --------------------------------------------------------
    # ZV-NUMMERN: FÜHRENDE NULL WIEDERHERSTELLEN
    # --------------------------------------------------------
    #
    # Im MIT1-Strangschema liefert die PDF-Textextraktion z. B.
    # "ZV70.1", obwohl die Positionsnummer im Plan als "ZV70.01"
    # geführt wird. Die führende Null ist im extrahierten PDF-Text
    # selbst nicht vorhanden und kann deshalb nicht einfach
    # "erhalten" werden.
    #
    # Für ZV-Positionsnummern ist der Teil nach dem ersten Punkt
    # zweistellig. Deshalb wird nur bei ZV und nur bei genau einer
    # Ziffer automatisch eine führende Null ergänzt:
    #
    #   ZV70.1   -> ZV70.01
    #   ZV70.9   -> ZV70.09
    #   ZV70.10  -> ZV70.10
    #   ZV70.12  -> ZV70.12
    #
    # Ein optionaler letzter Teil bleibt unverändert:
    #
    #   ZV70.1.2 -> ZV70.01.2
    #
    # Andere Präfixe werden NICHT aufgefüllt, weil Formate wie
    # P70.1 oder F1.2 laut allgemeiner Regel gültig sein können.
    parts = number.split(".")

    if (
        parts
        and parts[0].startswith("ZV")
        and len(parts) >= 2
        and len(parts[1]) == 1
        and parts[1].isdigit()
    ):
        parts[1] = parts[1].zfill(2)
        number = ".".join(parts)

    return number


def find_position_numbers(
    text: str,
) -> list[str]:
    """
    Findet alle Positionsnummern in einem Text.

    Wichtig:
    F24.527.1 und F24.527.2 werden als zwei verschiedene
    vollständige Nummern erkannt und nicht beide auf F24.527 gekürzt.
    """
    if not text:
        return []

    prepared = (
        str(
            text
        )
        .replace(
            "\u00a0",
            " ",
        )
        .replace(
            "\n",
            " ",
        )
        .replace(
            "\t",
            " ",
        )
        .upper()
    )

    numbers: list[str] = []

    for match in POSITION_PATTERN.finditer(
        prepared
    ):
        number = normalize_position_number(
            match.group(
                0
            )
        )

        if number is not None:
            numbers.append(
                number
            )

    return numbers


# ============================================================
# PDF-FARBEN
# ============================================================

def color_int_to_rgb(
    color: int,
) -> tuple[int, int, int]:
    """
    Wandelt eine PyMuPDF-Farbe in RGB um.
    """
    red = (
        color
        >> 16
    ) & 255

    green = (
        color
        >> 8
    ) & 255

    blue = (
        color
    ) & 255

    return (
        red,
        green,
        blue,
    )


def is_blueish(
    color: int,
) -> bool:
    """
    Prüft, ob eine Textfarbe blau/türkis ist.

    In den Schemata sind die relevanten Nummern meist
    türkis/blau.

    Beispiel PyMuPDF-Farbe:
        65535 = RGB(0, 255, 255)
    """
    red, green, blue = color_int_to_rgb(
        color
    )

    return (
        blue >= 120
        and green >= 80
        and red <= 120
    )


# ============================================================
# SCHEMA-PDF
# ============================================================

def _line_text_for_position_search(
    line: dict,
    only_blue: bool,
) -> str:
    """
    Baut den Suchtext einer PDF-Zeile aus ihren Spans zusammen.

    Das ist wichtig, weil eine Positionsnummer im PDF intern auf
    mehrere Spans verteilt sein kann, z. B.:

        "ZV" | "70" | "." | "01"

    Die alte spanweise Suche konnte solche Nummern übersehen.
    Jetzt wird daraus:

        "ZV 70 . 01"

    und das tolerante Regex erkennt korrekt ZV70.01.
    """
    pieces: list[str] = []

    for span in line.get(
        "spans",
        [],
    ):
        text = str(
            span.get(
                "text",
                "",
            )
        )

        if not text.strip():
            continue

        color = int(
            span.get(
                "color",
                0,
            )
        )

        if (
            only_blue
            and not is_blueish(
                color
            )
        ):
            continue

        pieces.append(
            text
        )

    return " ".join(
        pieces
    )


def extract_schema_positions(
    pdf_path: str | Path,
    only_blue: bool = True,
) -> pd.DataFrame:
    """
    Extrahiert Positionsnummern aus einem Schema-PDF.

    Verbesserung gegenüber der alten Version:
    Die Suche erfolgt zeilenweise über alle relevanten blauen Spans
    gemeinsam. Dadurch werden auch aufgesplittete Nummern wie
    ZV70.01 bis ZV70.12 zuverlässiger erkannt.
    """
    pdf_path = Path(
        pdf_path
    )

    records: list[
        PositionRecord
    ] = []

    document = fitz.open(
        pdf_path
    )

    try:
        for page_index, page in enumerate(
            document,
            start=1,
        ):
            page_dict = page.get_text(
                "dict"
            )

            for block in page_dict.get(
                "blocks",
                [],
            ):
                for line in block.get(
                    "lines",
                    [],
                ):
                    line_text = (
                        _line_text_for_position_search(
                            line,
                            only_blue=only_blue,
                        )
                    )

                    positions = find_position_numbers(
                        line_text
                    )

                    for position in positions:
                        records.append(
                            PositionRecord(
                                positionsnummer=position,
                                quelle="Schema",
                                datei=pdf_path.name,
                                seite=page_index,
                            )
                        )

    finally:
        document.close()

    return pd.DataFrame(
        [
            record.__dict__
            for record in records
        ],
        columns=[
            "positionsnummer",
            "quelle",
            "datei",
            "seite",
        ],
    )


# ============================================================
# BML-EXCEL
# ============================================================

def find_header_row_and_position_column(
    excel_path: str | Path,
    sheet_name: str = "Tabelle1",
) -> tuple[int, int]:
    """
    Findet die Header-Zeile und die Spalte 'Pos. Nr.'.

    Rückgabe:
    - header_row: 1-basierte Excel-Zeilennummer
    - position_column: 1-basierte Excel-Spaltennummer
    """
    workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Das Tabellenblatt '{sheet_name}' wurde "
                "in der Excel-Datei nicht gefunden."
            )

        worksheet = workbook[
            sheet_name
        ]

        for row in range(
            1,
            min(
                worksheet.max_row,
                30,
            )
            + 1,
        ):
            for column in range(
                1,
                worksheet.max_column
                + 1,
            ):
                value = worksheet.cell(
                    row=row,
                    column=column,
                ).value

                if value is None:
                    continue

                normalized = (
                    str(
                        value
                    )
                    .strip()
                    .lower()
                    .replace(
                        " ",
                        "",
                    )
                )

                if normalized in {
                    "pos.nr.",
                    "pos.nr",
                    "posnr.",
                    "posnr",
                }:
                    return (
                        row,
                        column,
                    )

    finally:
        workbook.close()

    raise ValueError(
        "Die Spalte 'Pos. Nr.' wurde in der Excel-Datei nicht gefunden."
    )


def extract_excel_positions(
    excel_path: str | Path,
    sheet_name: str = "Tabelle1",
) -> pd.DataFrame:
    """
    Extrahiert Positionsnummern aus der BML-Exceldatei.
    """
    excel_path = Path(
        excel_path
    )

    (
        header_row,
        position_column,
    ) = find_header_row_and_position_column(
        excel_path=excel_path,
        sheet_name=sheet_name,
    )

    dataframe = pd.read_excel(
        excel_path,
        sheet_name=sheet_name,
        header=header_row - 1,
    )

    position_column_name = (
        dataframe.columns[
            position_column - 1
        ]
    )

    records: list[
        PositionRecord
    ] = []

    for value in dataframe[
        position_column_name
    ].dropna():
        position = normalize_position_number(
            value
        )

        if position is None:
            continue

        records.append(
            PositionRecord(
                positionsnummer=position,
                quelle="Excel",
                datei=excel_path.name,
                seite=None,
            )
        )

    return pd.DataFrame(
        [
            record.__dict__
            for record in records
        ],
        columns=[
            "positionsnummer",
            "quelle",
            "datei",
            "seite",
        ],
    )


# ============================================================
# VERGLEICH
# ============================================================

def determine_status(
    schema_count: int,
    excel_count: int,
) -> str:
    """
    Bestimmt den Vergleichsstatus.
    """
    if (
        schema_count > 1
        and excel_count > 1
    ):
        return (
            "Mehrfach in beiden"
        )

    if schema_count > 1:
        return (
            "Mehrfach im Schema"
        )

    if excel_count > 1:
        return (
            "Mehrfach in Excel"
        )

    if (
        schema_count == 1
        and excel_count == 1
    ):
        return "OK"

    if (
        schema_count == 1
        and excel_count == 0
    ):
        return (
            "Nur im Schema"
        )

    if (
        schema_count == 0
        and excel_count == 1
    ):
        return (
            "Nur in Excel"
        )

    return "Unklar"


def compare_positions(
    schema_df: pd.DataFrame,
    excel_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Vergleicht Positionsnummern aus Schema und Excel.
    """
    schema_numbers = (
        schema_df[
            "positionsnummer"
        ].tolist()
        if not schema_df.empty
        else []
    )

    excel_numbers = (
        excel_df[
            "positionsnummer"
        ].tolist()
        if not excel_df.empty
        else []
    )

    schema_counter = Counter(
        schema_numbers
    )

    excel_counter = Counter(
        excel_numbers
    )

    all_numbers = sorted(
        set(
            schema_counter
        )
        | set(
            excel_counter
        )
    )

    rows: list[
        dict[str, object]
    ] = []

    for number in all_numbers:
        schema_count = schema_counter.get(
            number,
            0,
        )

        excel_count = excel_counter.get(
            number,
            0,
        )

        rows.append(
            {
                "Positionsnummer":
                    number,

                "Status":
                    determine_status(
                        schema_count,
                        excel_count,
                    ),

                "Anzahl Schema":
                    schema_count,

                "Anzahl Excel":
                    excel_count,
            }
        )

    result = pd.DataFrame(
        rows,
        columns=[
            "Positionsnummer",
            "Status",
            "Anzahl Schema",
            "Anzahl Excel",
        ],
    )

    if result.empty:
        return result

    status_order = {
        "Nur im Schema": 0,
        "Nur in Excel": 1,
        "Mehrfach im Schema": 2,
        "Mehrfach in Excel": 3,
        "Mehrfach in beiden": 4,
        "Unklar": 5,
        "OK": 6,
    }

    result[
        "_sort"
    ] = result[
        "Status"
    ].map(
        status_order
    )

    result = (
        result.sort_values(
            by=[
                "_sort",
                "Positionsnummer",
            ]
        )
        .drop(
            columns=[
                "_sort",
            ]
        )
        .reset_index(
            drop=True
        )
    )

    return result


# ============================================================
# EXCEL-AUSGABE
# ============================================================

def export_result(
    comparison_df: pd.DataFrame,
    schema_df: pd.DataFrame,
    excel_df: pd.DataFrame,
    output_path: str | Path,
) -> None:
    """
    Exportiert den Vergleich als Excel-Datei.

    Die Arbeitsmappe wird direkt mit openpyxl aufgebaut.
    Dadurch existiert zu jedem Zeitpunkt mindestens ein sichtbares
    Tabellenblatt. Das verhindert insbesondere in der gepackten EXE
    den openpyxl-Fehler "At least one sheet must be visible", der bei
    pandas.ExcelWriter eine ursprünglichere Export-Fehlermeldung
    überdecken kann.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    # Das von openpyxl automatisch angelegte sichtbare Blatt wird
    # direkt als erstes Ergebnisblatt verwendet.
    first_sheet = workbook.active
    first_sheet.title = "Vergleich"

    sheets = [
        ("Vergleich", comparison_df),
        ("Nummern_Schema", schema_df),
        ("Nummern_Excel", excel_df),
    ]

    for index, (sheet_name, dataframe) in enumerate(sheets):
        if index == 0:
            worksheet = first_sheet
        else:
            worksheet = workbook.create_sheet(
                title=sheet_name
            )

        # Auch leere DataFrames erhalten mindestens die Kopfzeile.
        for row in dataframe_to_rows(
            dataframe,
            index=False,
            header=True,
        ):
            worksheet.append(
                list(row)
            )

        # Falls ein DataFrame ausnahmsweise gar keine Spalten hat,
        # bleibt trotzdem ein sichtbares Blatt bestehen.
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            if worksheet["A1"].value is None:
                worksheet["A1"] = "Keine Daten"

    workbook.save(
        output_path
    )
    workbook.close()

    format_excel_output(
        output_path
    )

def format_excel_output(
    output_path: str | Path,
) -> None:
    """
    Formatiert die Excel-Auswertung.
    """
    workbook = load_workbook(
        output_path
    )

    status_colors = {
        "OK": "C6EFCE",
        "Nur im Schema": "FFF2CC",
        "Nur in Excel": "F4CCCC",
        "Mehrfach im Schema": "FCE4D6",
        "Mehrfach in Excel": "FCE4D6",
        "Mehrfach in beiden": "EADCF8",
        "Unklar": "D9EAD3",
    }

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = (
            "A2"
        )

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[
            1
        ]:
            cell.font = Font(
                bold=True
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )

        for column_cells in worksheet.columns:
            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[
                        0
                    ].column
                )
            )

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(
                            str(
                                cell.value
                            )
                        ),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                45,
            )

    comparison_sheet = workbook[
        "Vergleich"
    ]

    status_column = None

    for cell in comparison_sheet[
        1
    ]:
        if cell.value == "Status":
            status_column = (
                cell.column
            )
            break

    if status_column is not None:
        for row in range(
            2,
            comparison_sheet.max_row
            + 1,
        ):
            status = (
                comparison_sheet.cell(
                    row=row,
                    column=status_column,
                ).value
            )

            color = status_colors.get(
                status
            )

            if color:
                for column in range(
                    1,
                    comparison_sheet.max_column
                    + 1,
                ):
                    comparison_sheet.cell(
                        row=row,
                        column=column,
                    ).fill = PatternFill(
                        fill_type="solid",
                        fgColor=color,
                    )

    workbook.save(
        output_path
    )


# ============================================================
# HAUPTFUNKTION
# ============================================================

def run_hk_number_check(
    schema_pdf: str | Path,
    bml_excel: str | Path,
    output_dir: str | Path,
    name: str = "HK_Nummernkontrolle",
) -> Path:
    """
    Führt die komplette Nummernkontrolle
    für ein PDF-Excel-Paar aus.
    """
    schema_pdf = Path(
        schema_pdf
    )

    bml_excel = Path(
        bml_excel
    )

    output_dir = Path(
        output_dir
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    schema_df = (
        extract_schema_positions(
            schema_pdf,
            only_blue=True,
        )
    )

    excel_df = (
        extract_excel_positions(
            bml_excel
        )
    )

    comparison_df = (
        compare_positions(
            schema_df,
            excel_df,
        )
    )

    safe_name = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        name,
    ).strip(
        "_"
    )

    output_path = (
        output_dir
        / f"{safe_name}.xlsx"
    )

    export_result(
        comparison_df=comparison_df,
        schema_df=schema_df,
        excel_df=excel_df,
        output_path=output_path,
    )

    return output_path
