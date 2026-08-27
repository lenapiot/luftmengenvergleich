from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook


LoadMode = Literal["heizung", "kuehlung", "beides"]


ROOM_HEADER = "Nummer"
HEATING_HEADER = "Bemessungslast Heizung"
COOLING_HEADER = "Bemessungslast Kühlung"


@dataclass(frozen=True)
class ExcelLoadRecord:
    """
    Ein aus der Heiz-/Kühllast-Excel gelesener Raum.

    raum:
        Originale Raumnummer aus Excel, z. B. MIT1J117a.

    raum_key:
        Normalisierte Vergleichsform, z. B. MIT1J117A.

    heizlast_w / kuehllast_w:
        Vergleichswert als positiver Betrag in Watt.
        Nicht ausgewählte Lastarten sind None.

    heizlast_original_w / kuehllast_original_w:
        Originalwert aus Excel inklusive Vorzeichen.

    excel_zeile:
        Ursprüngliche Zeilennummer in Excel.
    """

    raum: str
    raum_key: str
    heizlast_w: int | None
    kuehllast_w: int | None
    heizlast_original_w: int | None
    kuehllast_original_w: int | None
    excel_zeile: int


def _normalize_header(value: object) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\u00a0", " ").strip(),
    ).casefold()


def normalize_room_key(value: object) -> str | None:
    """
    Normalisiert Raumnummern für den späteren Vergleich.

    Beispiele:
        MIT1J111   -> MIT1J111
        MIT1J117a  -> MIT1J117A
        ' MIT1 H 320a ' -> MIT1H320A

    Es werden nur Leerzeichen entfernt und Gross-/Kleinschreibung
    vereinheitlicht. Die eigentliche Nummer wird nicht umgebaut.
    """
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    compact = re.sub(r"\s+", "", text).upper()

    # Bewusst relativ tolerant:
    # MIT1J111, MIT1J117A, MIT1H320A, MIT2...
    if not re.fullmatch(r"MIT\d+[A-Z]+\d+[A-Z]?", compact):
        return None

    return compact


def parse_load_w(value: object) -> int | None:
    """
    Liest Leistungswerte in Watt.

    Unterstützt u. a.:
        539 W
        -539 W
        0 W
        539
        539.0

    Tausendertrennzeichen und geschützte Leerzeichen werden entfernt.
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return int(round(float(value)))

    text = (
        str(value)
        .replace("\u00a0", " ")
        .replace("'", "")
        .replace("’", "")
        .strip()
    )

    if not text:
        return None

    # Einheit entfernen.
    text = re.sub(
        r"\s*[wW]\s*$",
        "",
        text,
    ).strip()

    # Normale Leerzeichen als Tausendertrennzeichen entfernen.
    text = text.replace(" ", "")

    # Komma als Dezimaltrennzeichen unterstützen.
    text = text.replace(",", ".")

    try:
        return int(round(float(text)))
    except ValueError:
        return None


def _find_columns(
    worksheet,
    max_header_rows: int = 20,
) -> tuple[int, int, int, int]:
    """
    Sucht Kopfzeile und die drei Pflichtspalten:
        Nummer
        Bemessungslast Heizung
        Bemessungslast Kühlung
    """
    wanted = {
        _normalize_header(ROOM_HEADER): "room",
        _normalize_header(HEATING_HEADER): "heating",
        _normalize_header(COOLING_HEADER): "cooling",
    }

    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=(
                min(
                    worksheet.max_row,
                    max_header_rows,
                )
                if (
                    worksheet.max_row is not None
                    and worksheet.max_row >= 1
                )
                else max_header_rows
            ),
            values_only=True,
        ),
        start=1,
    ):
        found: dict[str, int] = {}

        for col_index, value in enumerate(
            row,
            start=1,
        ):
            normalized = _normalize_header(value)

            key = wanted.get(normalized)

            if key is not None:
                found[key] = col_index

        if {
            "room",
            "heating",
            "cooling",
        }.issubset(found):
            return (
                row_index,
                found["room"],
                found["heating"],
                found["cooling"],
            )

    raise ValueError(
        "Die benötigten Spalten wurden nicht gefunden.\n\n"
        "Erwartet werden:\n"
        f"- {ROOM_HEADER}\n"
        f"- {HEATING_HEADER}\n"
        f"- {COOLING_HEADER}"
    )


def _select_worksheet(workbook):
    """
    Nimmt das erste Tabellenblatt, das die drei benötigten Spalten enthält.
    Dadurch hängt die Extraktion nicht von einem bestimmten Blattnamen ab.
    """
    errors: list[str] = []

    for worksheet in workbook.worksheets:
        try:
            columns = _find_columns(worksheet)
            return worksheet, columns

        except ValueError:
            errors.append(worksheet.title)

    raise ValueError(
        "In keinem Tabellenblatt wurden die benötigten Spalten gefunden.\n\n"
        "Geprüfte Tabellenblätter:\n- "
        + "\n- ".join(errors)
    )


def extract_loads_from_excel(
    excel_path: str | Path,
    mode: LoadMode = "beides",
) -> list[ExcelLoadRecord]:
    """
    Extrahiert Heiz- und/oder Kühllasten aus der Raummodell-Excel.

    mode:
        "heizung"   -> nur Heizlast wird übernommen
        "kuehlung"  -> nur Kühllast wird übernommen
        "beides"    -> beide Lastarten werden übernommen

    Kühlwerte werden als positive Leistung gespeichert:
        -539 W -> 539 W

    Heizwerte werden ebenfalls als Betrag gespeichert.
    Dadurch ist die Einheit/Logik für den späteren Vergleich einheitlich.
    """
    if mode not in {
        "heizung",
        "kuehlung",
        "beides",
    }:
        raise ValueError(
            "Ungültiger Modus. Erlaubt sind: "
            "'heizung', 'kuehlung', 'beides'."
        )

    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel-Datei nicht gefunden: {excel_path}"
        )

    if excel_path.suffix.lower() not in {
        ".xlsx",
        ".xlsm",
    }:
        raise ValueError(
            "Die Lastdatei muss eine .xlsx- oder .xlsm-Datei sein."
        )

    workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet, (
            header_row,
            room_col,
            heating_col,
            cooling_col,
        ) = _select_worksheet(workbook)

        records: list[ExcelLoadRecord] = []

        min_col = min(
            room_col,
            heating_col,
            cooling_col,
        )
        max_col = max(
            room_col,
            heating_col,
            cooling_col,
        )

        for excel_row, values in enumerate(
            worksheet.iter_rows(
                min_row=header_row + 1,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            def local_value(column: int):
                return values[
                    column - min_col
                ]

            room_raw = local_value(
                room_col
            )

            room_key = normalize_room_key(
                room_raw
            )

            if room_key is None:
                # Leere / ungültige Zeilen werden ignoriert.
                continue

            heating_raw = parse_load_w(
                local_value(
                    heating_col
                )
            )
            cooling_raw = parse_load_w(
                local_value(
                    cooling_col
                )
            )

            if mode in {
                "heizung",
                "beides",
            }:
                heating_value = (
                    abs(heating_raw)
                    if heating_raw is not None
                    else None
                )
            else:
                heating_value = None

            if mode in {
                "kuehlung",
                "beides",
            }:
                cooling_value = (
                    abs(cooling_raw)
                    if cooling_raw is not None
                    else None
                )
            else:
                cooling_value = None

            records.append(
                ExcelLoadRecord(
                    raum=str(room_raw).strip(),
                    raum_key=room_key,
                    heizlast_w=heating_value,
                    kuehllast_w=cooling_value,
                    heizlast_original_w=(
                        heating_raw
                        if mode in {"heizung", "beides"}
                        else None
                    ),
                    kuehllast_original_w=(
                        cooling_raw
                        if mode in {"kuehlung", "beides"}
                        else None
                    ),
                    excel_zeile=excel_row,
                )
            )

        if not records:
            raise ValueError(
                "Es wurden keine gültigen Raum-/Lastdaten in der Excel-Datei gefunden."
            )

        return records

    finally:
        workbook.close()


def summarize_excel_records(
    records: list[ExcelLoadRecord],
) -> dict[str, int]:
    """
    Kleine Zusammenfassung für Tests und spätere GUI-Rückmeldungen.
    """
    return {
        "anzahl_raeume": len(records),
        "mit_heizlast": sum(
            record.heizlast_w is not None
            for record in records
        ),
        "mit_kuehllast": sum(
            record.kuehllast_w is not None
            for record in records
        ),
        "heizlast_0w": sum(
            record.heizlast_w == 0
            for record in records
            if record.heizlast_w is not None
        ),
        "kuehllast_0w": sum(
            record.kuehllast_w == 0
            for record in records
            if record.kuehllast_w is not None
        ),
    }
